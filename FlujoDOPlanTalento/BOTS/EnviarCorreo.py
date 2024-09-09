from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
import smtplib
import os
import mimetypes
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

InsumoDesarrolloHumano = os.getenv('InsumoDesarrolloHumano')
PDFCarpeta = os.getenv('PDFCarpeta')

remitente = os.getenv('SMTP_USERNAME')
contrasena = os.getenv('SMTP_PASSWORD')
servidor_smtp = os.getenv('SMTP_SERVER')
puerto_smtp = int(os.getenv('SMTP_PORT'))


ruta_imagen_cierreBrechas = os.getenv('ruta_imagen_cierreBrechas')
ruta_imagen_Acompanamiento = os.getenv('ruta_imagen_Acompanamiento')

def enviar_correo_CierreBrechas(destinatario, ruta_pdf, destinatarios_bcc=None):
    # Crear el objeto del mensaje
    msg = MIMEMultipart()
    msg['From'] = remitente
    msg['To'] = destinatario
    msg['Subject'] = "Cierre de brechas Plan Talento."
    ruta_imagen = ruta_imagen_cierreBrechas
    mensaje_html = """
        <html>
        <body>
            <p style="text-align: center;">
            <img src="cid:imagen1" style="width: 600px; height: auto;">
            </p>
        </body>
        </html>
        """
    
    # Adjuntar el cuerpo del mensaje en HTML
    msg.attach(MIMEText(mensaje_html, 'html'))

    with open(ruta_imagen, 'rb') as f:
        img_data = f.read()
    # Determinar el tipo MIME de la imagen
    mime_type, _ = mimetypes.guess_type(ruta_imagen)
    if mime_type is None:
        raise TypeError('Could not guess image MIME subtype')
    main_type, sub_type = mime_type.split('/')
    img = MIMEImage(img_data, _subtype=sub_type)
    img.add_header('Content-ID', '<imagen1>')
    img.add_header('Content-Disposition', 'inline', filename=os.path.basename(ruta_imagen))
    msg.attach(img)
    
    # Adjuntar el PDF
    with open(ruta_pdf, 'rb') as f:
        pdf_data = f.read()
    pdf = MIMEApplication(pdf_data, _subtype='pdf')
    pdf.add_header('Content-Disposition', 'attachment', filename=os.path.basename(ruta_pdf))
    msg.attach(pdf)

    try:
        # Establecer conexión con el servidor SMTP
        servidor = smtplib.SMTP(servidor_smtp, puerto_smtp)
        servidor.starttls()
        servidor.login(remitente, contrasena)

        # Enviar el mensaje
        if destinatarios_bcc:
            servidor.sendmail(remitente, [destinatario] + destinatarios_bcc, msg.as_string())
        else:
            servidor.sendmail(remitente, destinatario, msg.as_string())

        # Cerrar la conexión
        servidor.quit()

        print("Correo enviado exitosamente Cierre de Brechas")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

def enviar_correo_Acompanamiento(destinatario, ruta_pdf):
    # Crear el objeto del mensaje
    msg = MIMEMultipart()
    msg['From'] = remitente
    msg['To'] = destinatario
    msg['Subject'] = "Acompañamiento cierre de brechas Plan Talento."
    ruta_imagen = ruta_imagen_Acompanamiento
    mensaje_html = """
        <html>
        <body>
            <p style="text-align: center;">
            <img src="cid:imagen1" style="width: 600px; height: auto;">
            </p>
        </body>
        </html>
        """
    
    # Adjuntar el cuerpo del mensaje en HTML
    msg.attach(MIMEText(mensaje_html, 'html'))

    with open(ruta_imagen, 'rb') as f:
        img_data = f.read()
    # Determinar el tipo MIME de la imagen
    mime_type, _ = mimetypes.guess_type(ruta_imagen)
    if mime_type is None:
        raise TypeError('Could not guess image MIME subtype')
    main_type, sub_type = mime_type.split('/')
    img = MIMEImage(img_data, _subtype=sub_type)
    img.add_header('Content-ID', '<imagen1>')
    img.add_header('Content-Disposition', 'inline', filename=os.path.basename(ruta_imagen))
    msg.attach(img)
    
    # Adjuntar el PDF
    with open(ruta_pdf, 'rb') as f:
        pdf_data = f.read()
    pdf = MIMEApplication(pdf_data, _subtype='pdf')
    pdf.add_header('Content-Disposition', 'attachment', filename=os.path.basename(ruta_pdf))
    msg.attach(pdf)

    try:
        # Establecer conexión con el servidor SMTP
        servidor = smtplib.SMTP(servidor_smtp, puerto_smtp)
        servidor.starttls()
        servidor.login(remitente, contrasena)

        # Enviar el mensaje
        servidor.sendmail(remitente, destinatario, msg.as_string())

        # Cerrar la conexión
        servidor.quit()

        print("Correo enviado exitosamente Acompañamiento")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

# Cargar el archivo Excel usando openpyxl
wb = load_workbook(InsumoDesarrolloHumano)
ws = wb.active

# Identificar los encabezados de las columnas
headers = {cell.value: cell.column for cell in ws[1]}

# Filtrar filas basadas en las condiciones y extraer la información requerida
filas_filtradas = []
for row in ws.iter_rows(min_row=2):
    if row[headers['Habilitado'] - 1].value == 'ABIERTO' and row[headers['Enviado'] - 1].value == 'PENDIENTE' and row[headers['Automatizacion'] - 1].value == 'Excel':
        filas_filtradas.append(row)

# Verificar si hay filas que cumplan con las condiciones
if filas_filtradas:
    archivos_pdf = os.listdir(PDFCarpeta)
    celdas_con_pdf = []
    
    for fila in filas_filtradas:
        cedula = str(fila[headers['Cédula'] - 1].value)  # Convertir a cadena para comparación
        nombre_pdf = f"{cedula}.pdf"
        if nombre_pdf in archivos_pdf:
            print(f"Existe un PDF correspondiente para la cédula: {cedula}")
            ruta_pdf = os.path.join(PDFCarpeta, nombre_pdf)
            
            bcc = ['luisa.quintana@gruporeditos.com', 'luisa.torres@gruporeditos.com']
            
            # Toma el correo de la Base de Datos del Postulante
            destinatario_postulante = fila[headers['Correo electrónico '] - 1].value.strip()
            enviar_correo_CierreBrechas(destinatario_postulante, ruta_pdf, destinatarios_bcc=bcc)
            
            # Toma el correo del Nuevo Lider
            destinatario_nuevolider = fila[headers['Correo Nuevo Lider'] - 1].value.strip()
            enviar_correo_Acompanamiento(destinatario_nuevolider, ruta_pdf)
            
            celdas_con_pdf.append(cedula)

            # Actualizar las columnas correspondientes
            fila[headers['Estado'] - 1].value = 'COMPLETO'
            fila[headers['Habilitado'] - 1].value = 'CERRADO'
            fila[headers['Enviado'] - 1].value = 'ENVIADO'
            fila[headers['Automatizacion'] - 1].value = 'PDF'
            
            # Eliminar el archivo PDF
            try:
                os.remove(ruta_pdf)
                print(f"El archivo {nombre_pdf} ha sido eliminado.")
            except Exception as e:
                print(f"Error al eliminar el archivo {nombre_pdf}: {e}")
        else:
            print(f"No se encontró el PDF para la cédula: {cedula}.")

else:
    print("No hay filas que cumplan con las condiciones especificadas.")

# Guardar los cambios en el archivo Excel
wb.save(InsumoDesarrolloHumano)

