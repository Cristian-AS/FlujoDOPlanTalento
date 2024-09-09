# Version: 1.0
""""
Este script se encarga de leer un archivo Excel que contiene información sobre el desarrollo humano de los empleados. 
Luego, filtra las filas que contienen la palabra 'excel' en la columna 'Automatizacion'. Por cada fila filtrada, se busca un archivo 
Excel en una carpeta que contiene información sobre el cambio de cargo de un empleado. Si el archivo existe, se abre y se pega la 
información de competencias desarrolladas en las primeras filas y columnas de la hoja 'Competencias Desarrolladas'. Finalmente, 
se guarda el archivo Excel con la información actualizada. 
"""

# Importar librerías necesarias
import os
from dotenv import load_dotenv
import pandas as pd
import openpyxl

# Cargar las variables de entorno
load_dotenv()

# Definir las rutas de los archivos y carpetas
InsumoDesarrolloHumano = os.getenv('InsumoDesarrolloHumano')
InsumoCambioCargo = os.getenv('InsumoCambioCargo')

# Leer el archivo Excel principal
df = pd.read_excel(InsumoDesarrolloHumano)

# Filtrar las filas donde la columna 'Automatizacion' tiene el valor 'Excel'
filtered_df = df[df['Automatizacion'] == 'Excel']

# Iterar sobre cada fila del DataFrame filtrado
for index, row in filtered_df.iterrows():
    cedula = int(row['Cédula'])
    filename = f"{cedula}.xlsx"
    filepath = os.path.join(InsumoCambioCargo, filename)
    
    # Verificar si el archivo Excel existe
    if os.path.exists(filepath):
        print(f"Archivo encontrado para la cédula {cedula}: {filepath}")
        
        # Abrir el archivo Excel y cargar el libro de trabajo
        wb = openpyxl.load_workbook(filepath)
        
        # Verificar si la hoja "Competencias Desarrolladas" existe en el archivo Excel
        if 'Competencias Desarrolladas' in wb.sheetnames:
            sheet = wb['Competencias Desarrolladas']
            
            # Obtener las columnas de Competencia_Insumo correspondientes a esta fila
            competencia_insumo = [
                row['Plan Accion Experiencial'], row['Fecha Inicial Experiencial'], row['Fecha Final Experiencial'],
                row['Plan Accion Social'], row['Fecha Inicial Social'], row['Fecha Final Social'],
                row['Plan Accion Formal'], row['Fecha Inicial Formal'], row['Fecha Final Formal']
            ]
            
            # Pegar los valores en las primeras filas y columnas
            for col_idx, value in enumerate(competencia_insumo, start=1):
                sheet.cell(row=2, column=col_idx).value = value
                
            # Guardar el archivo Excel
            wb.save(filepath)
            wb.close()
            
            print(f"Información de competencias pegada en el archivo {filename}")
        else:
            print(f"La hoja 'Competencias Desarrolladas' no existe en el archivo {filename}")
        
    else:
        print(f"No existe archivo para la cédula {cedula}")


